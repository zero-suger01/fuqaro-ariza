"""Operatsion navbat — `stats/queues` va ro'yxat filtrlari
(docs/03-kontraktlar.md §5, docs/10-ui-ux.md §10.1).

Asosiy shart: **karta raqami va karta bosilganda ochiladigan ro'yxat
bir xil shartdan chiqadi**. Ular farq qilsa panelga ishonch qolmaydi,
shuning uchun har navbat uchun ikkisi ham tekshiriladi. Shu zanjirning
uchinchi bo'g'ini — `export.xlsx` (pastdagi oxirgi ikki test).
"""
import io
import uuid
from datetime import datetime, timedelta, timezone

import pytest
from openpyxl import load_workbook

from app.models.complaint import Complaint

# Frontend'dagi `QUEUES` xaritasi (admin/murojaatlar/page.tsx) yuboradigan
# parametrlar — ro'yxat va eksport ikkisi ham shu to'plamni tushunishi kerak.
QUEUE_PARAMS = {
    "unassigned": {"unassigned": True},
    "sla_risk": {"sla_risk": True},
    "overdue": {"overdue": True},
    "need_info": {"need_info_over_hours": 24},
    "ai": {"needs_review": True},
    "stuck_ai": {"stuck_ai": True},
    "mine": {"mine": True},
}


def _submit(client, description: str = "Ko'chada chiroq yonmayapti, kechasi qorong'i") -> tuple[str, str]:
    phone = f"+99895{uuid.uuid4().int % 10**7:07d}"
    response = client.post(
        "/api/public/complaints",
        data={"description": description, "first_name": "Test", "phone": phone},
    )
    assert response.status_code == 201, response.text
    return response.json()["ticket_number"], phone


def _find(db_session, ticket: str) -> Complaint:
    return db_session.query(Complaint).filter(Complaint.ticket_number == ticket).one()


def _queues(client, admin_headers) -> dict:
    response = client.get("/api/admin/stats/queues", headers=admin_headers)
    assert response.status_code == 200, response.text
    return response.json()


def _list_ids(client, admin_headers, **params) -> set[str]:
    response = client.get(
        "/api/admin/complaints", params={**params, "page_size": 100}, headers=admin_headers
    )
    assert response.status_code == 200, response.text
    return {item["id"] for item in response.json()["items"]}


def _list_tickets(client, admin_headers, **params) -> list[str]:
    """Ro'yxatdagi ticketlar — TARTIBI bilan (eksport tartibi bilan
    solishtirish uchun)."""
    response = client.get(
        "/api/admin/complaints", params={**params, "page_size": 100}, headers=admin_headers
    )
    assert response.status_code == 200, response.text
    return [item["ticket_number"] for item in response.json()["items"]]


def _export_tickets(client, admin_headers, **params) -> list[str]:
    """Eksport faylining 1-ustuni (`Ticket`) — sarlavha qatorisiz."""
    response = client.get("/api/admin/complaints/export.xlsx", params=params, headers=admin_headers)
    assert response.status_code == 200, response.text
    ws = load_workbook(io.BytesIO(response.content)).active
    return [row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True)]


@pytest.mark.smoke
def test_queues_shape_and_departments(client, admin_headers):
    body = _queues(client, admin_headers)
    for key in ("unassigned", "ai_exceptions", "sla_risk", "overdue", "awaiting_info", "stuck_ai"):
        assert isinstance(body[key], int), key
    assert body["by_department"], "seed'da faol bo'limlar bo'lishi kerak"
    row = body["by_department"][0]
    for key in ("department_name", "new", "in_progress", "sla_risk", "overdue", "unowned", "over_limit"):
        assert key in row, key


@pytest.mark.smoke
def test_new_complaint_counts_as_unassigned(client, admin_headers, db_session):
    before = _queues(client, admin_headers)["unassigned"]
    ticket, _ = _submit(client)
    after = _queues(client, admin_headers)

    assert after["unassigned"] == before + 1
    assert str(_find(db_session, ticket).id) in _list_ids(client, admin_headers, unassigned=True)


@pytest.mark.smoke
def test_overdue_card_matches_the_list(client, admin_headers, db_session):
    ticket, _ = _submit(client)
    complaint = _find(db_session, ticket)

    before = _queues(client, admin_headers)["overdue"]
    complaint.deadline_at = datetime.now(timezone.utc) - timedelta(hours=1)
    db_session.flush()

    after = _queues(client, admin_headers)
    assert after["overdue"] == before + 1
    assert str(complaint.id) in _list_ids(client, admin_headers, overdue=True)
    # Muddati o'tgan ish SLA-xavf navbatida takrorlanmaydi.
    assert str(complaint.id) not in _list_ids(client, admin_headers, sla_risk=True)


@pytest.mark.smoke
def test_sla_risk_uses_the_75_percent_threshold(client, admin_headers, db_session):
    """Chegara `escalation.py` dagi `sla_warning` bilan bir xil — xabar
    olgan xodim ishni navbatda topishi kerak."""
    ticket, _ = _submit(client)
    complaint = _find(db_session, ticket)
    now = datetime.now(timezone.utc)

    # 50% sarflangan — hali xavf emas.
    complaint.created_at = now - timedelta(hours=5)
    complaint.deadline_at = now + timedelta(hours=5)
    db_session.flush()
    assert str(complaint.id) not in _list_ids(client, admin_headers, sla_risk=True)

    # 90% sarflangan — xavf.
    complaint.created_at = now - timedelta(hours=9)
    complaint.deadline_at = now + timedelta(hours=1)
    db_session.flush()

    assert str(complaint.id) in _list_ids(client, admin_headers, sla_risk=True)
    assert _queues(client, admin_headers)["sla_risk"] >= 1


@pytest.mark.smoke
def test_stuck_ai_catches_silent_llm(client, admin_headers, db_session):
    """`status=new` da 1 soatdan ko'p turgan murojaat — LLM javob
    bermagani (docs/07 §2.3). Sukut bo'yicha bu navbat bo'sh bo'lishi
    kerak, ichida yozuv paydo bo'lishi uzilish signali."""
    ticket, _ = _submit(client)
    complaint = _find(db_session, ticket)
    assert complaint.status == "new"

    assert str(complaint.id) not in _list_ids(client, admin_headers, stuck_ai=True)

    complaint.created_at = datetime.now(timezone.utc) - timedelta(hours=2)
    db_session.flush()

    assert str(complaint.id) in _list_ids(client, admin_headers, stuck_ai=True)
    assert _queues(client, admin_headers)["stuck_ai"] >= 1


@pytest.mark.smoke
def test_awaiting_info_needs_24_hours(client, admin_headers, db_session):
    ticket, _ = _submit(client)
    complaint = _find(db_session, ticket)
    complaint_id = str(complaint.id)

    client.patch(f"/api/admin/complaints/{complaint_id}/status", json={"status": "ai_processed"}, headers=admin_headers)
    client.post(f"/api/admin/complaints/{complaint_id}/review", json={"reason": "ok"}, headers=admin_headers)
    asked = client.patch(
        f"/api/admin/complaints/{complaint_id}/status",
        json={"status": "need_info", "note": "Aniq manzilni yozing"},
        headers=admin_headers,
    )
    assert asked.status_code == 200, asked.text

    # Hozir so'ralgan — hali follow-up vaqti emas.
    assert complaint_id not in _list_ids(client, admin_headers, need_info_over_hours=24)

    complaint.info_requested_at = datetime.now(timezone.utc) - timedelta(hours=30)
    db_session.flush()

    assert complaint_id in _list_ids(client, admin_headers, need_info_over_hours=24)
    assert _queues(client, admin_headers)["awaiting_info"] >= 1


@pytest.mark.smoke
def test_wip_limit_flags_but_does_not_block(client, admin_headers, db_session):
    """`wip_limit` faqat ko'rsatkich — hech narsani bloklamaydi ([04])."""
    from app.models.department import Department

    ticket, _ = _submit(client)
    complaint_id = str(_find(db_session, ticket).id)
    client.patch(f"/api/admin/complaints/{complaint_id}/status", json={"status": "ai_processed"}, headers=admin_headers)
    assigned = client.post(
        f"/api/admin/complaints/{complaint_id}/review", json={"reason": "ok"}, headers=admin_headers
    )
    department_id = assigned.json()["department"]["id"]

    department = db_session.get(Department, uuid.UUID(department_id))
    department.wip_limit = 0
    db_session.flush()

    row = next(r for r in _queues(client, admin_headers)["by_department"] if r["department_id"] == department_id)
    assert row["wip_limit"] == 0
    assert row["over_limit"] is True

    # Limit oshgan bo'lsa ham ish davom etadi.
    still_works = client.patch(
        f"/api/admin/complaints/{complaint_id}/status", json={"status": "in_progress"}, headers=admin_headers
    )
    assert still_works.status_code == 200, still_works.text


@pytest.mark.smoke
def test_subtask_blocks_resolve(client, admin_headers, db_session):
    """Ochiq idoralararo topshiriq bilan `resolved` — 422 `subtasks_open`."""
    ticket, _ = _submit(client)
    complaint_id = str(_find(db_session, ticket).id)
    client.patch(f"/api/admin/complaints/{complaint_id}/status", json={"status": "ai_processed"}, headers=admin_headers)
    client.post(f"/api/admin/complaints/{complaint_id}/review", json={"reason": "ok"}, headers=admin_headers)
    client.patch(f"/api/admin/complaints/{complaint_id}/status", json={"status": "in_progress"}, headers=admin_headers)

    departments = client.get("/api/admin/departments", headers=admin_headers).json()
    other = next(d for d in departments if d["code"] == "suvsoz")
    created = client.post(
        f"/api/admin/complaints/{complaint_id}/subtasks",
        json={"department_id": other["id"], "note": "Drenaj quvurini tekshirish"},
        headers=admin_headers,
    )
    assert created.status_code == 201, created.text
    subtask_id = created.json()["subtasks"][0]["id"]

    blocked = client.patch(
        f"/api/admin/complaints/{complaint_id}/status",
        json={"status": "resolved", "reply_text": "Hal qilindi."},
        headers=admin_headers,
    )
    assert blocked.status_code == 422, blocked.text
    assert blocked.json()["code"] == "subtasks_open"

    closed = client.patch(f"/api/admin/subtasks/{subtask_id}", json={"status": "done"}, headers=admin_headers)
    assert closed.status_code == 200, closed.text

    now_ok = client.patch(
        f"/api/admin/complaints/{complaint_id}/status",
        json={"status": "resolved", "reply_text": "Drenaj tozalandi, muammo hal bo'ldi."},
        headers=admin_headers,
    )
    assert now_ok.status_code == 200, now_ok.text


def _seed_one_per_queue(client, admin_headers, db_session) -> dict[str, str]:
    """Har navbatga kamida bitta murojaat — navbatlar bo'sh bo'lsa
    pastdagi solishtirish `set() == set()` ga aylanib, hech narsani
    tekshirmagan bo'lardi.

    Shartlar `services/queues.py` dagi maydonlarga to'g'ridan-to'g'ri
    qo'yiladi: bu yerda holat o'tishlari emas, ro'yxat va eksportning
    BIR XIL to'plam berishi tekshiriladi.
    """
    from app.models.department import Department

    now = datetime.now(timezone.utc)
    tickets: dict[str, str] = {}

    def seed(key: str, **fields) -> Complaint:
        ticket, _ = _submit(client)
        complaint = _find(db_session, ticket)
        for name, value in fields.items():
            setattr(complaint, name, value)
        tickets[key] = ticket
        return complaint

    seed("unassigned")  # yangi murojaat — bo'limi ham, xodimi ham yo'q
    seed("overdue", deadline_at=now - timedelta(hours=1))
    seed("sla_risk", created_at=now - timedelta(hours=9), deadline_at=now + timedelta(hours=1))
    seed("need_info", status="need_info", info_requested_at=now - timedelta(hours=30))
    seed("ai", needs_review=True)
    seed("stuck_ai", created_at=now - timedelta(hours=2))

    # `mine` — VA bir vaqtda «biriktirilgan» yozuv: bo'limi ham, xodimi
    # ham bor, ya'ni `unassigned` navbatidan tushadi. Shu bitta yozuv
    # tufayli quyidagi «navbat butun bazadan kichik» tekshiruvi haqiqiy
    # ma'no kasb etadi.
    me = client.get("/api/auth/me", headers=admin_headers)
    assert me.status_code == 200, me.text
    department_id = db_session.query(Department.id).filter(Department.is_active.is_(True)).first()[0]
    seed("mine", assigned_user_id=uuid.UUID(me.json()["id"]), assigned_department_id=department_id)

    db_session.flush()
    return tickets


@pytest.mark.smoke
def test_export_respects_every_queue_filter(client, admin_headers, db_session):
    """`export.xlsx` ekranda ko'rinayotgan AYNAN shu to'plamni beradi.

    QA'da topilgan xato: eksport endpoint'i `unassigned`, `sla_risk`,
    `need_info_over_hours`, `mine`, `stuck_ai` parametrlarini E'LON
    QILMAGANDI. FastAPI e'lon qilinmagan query parametrni indamay
    tashlab yuboradi — natijada «Biriktirilmagan» navbatidan (dev bazada
    98 ta) eksport qilinganda faylga BARCHA 119 murojaat tushardi va
    admin buni sezmasdi. Faqat `overdue` tasodifan ishlagan, chunki u
    e'lon qilingan yagona navbat parametri edi.
    """
    _seed_one_per_queue(client, admin_headers, db_session)
    everything = _export_tickets(client, admin_headers)

    for name, params in QUEUE_PARAMS.items():
        expected = _list_tickets(client, admin_headers, **params)
        assert expected, f"{name}: navbat bo'sh — tekshiruv ma'nosiz"

        exported = _export_tickets(client, admin_headers, **params)
        # Tartib ham teng: eksport ekrandagi navbat tartibida chiqadi ([03] v1.9).
        assert exported == expected, f"{name}: eksport ro'yxatdan farq qiladi"
        # Filtr HAQIQATAN qisqartirgan bo'lishi kerak — aks holda yuqoridagi
        # tenglik «hamma narsa == hamma narsa» bo'lib o'tib ketardi.
        assert len(exported) < len(everything), f"{name}: eksport butun bazani qaytardi"


@pytest.mark.smoke
def test_export_respects_stage_tabs(client, admin_headers, db_session):
    """Bosqich tabi (`stage=new|progress|done`) ham eksportga o'tishi kerak.

    v1.9 da `stage` ro'yxatga qo'shildi, eksport imzosiga esa o'sha zahoti
    qo'shilmadi — yuqoridagi v1.4 navbat bayroqlari bilan bo'lgan xatoning
    AYNAN o'zi (FastAPI e'lon qilinmagan query parametrni jimgina
    tashlaydi). Ya'ni bu drift ikki marta takrorlangan, shuning uchun
    alohida qo'riqchi test.
    """
    _seed_one_per_queue(client, admin_headers, db_session)
    everything = _export_tickets(client, admin_headers)

    seen: list[str] = []
    for stage in ("new", "progress", "done"):
        expected = _list_tickets(client, admin_headers, stage=stage)
        exported = _export_tickets(client, admin_headers, stage=stage)
        assert exported == expected, f"stage={stage}: eksport ro'yxatdan farq qiladi"
        seen.extend(exported)

    # Uchala bosqich birgalikda hammasini qoplaydi va kesishmaydi. Busiz
    # yuqoridagi tenglik bo'sh to'plamlarda ham «o'tib» ketardi; qo'shimcha
    # ravishda bu `constants.STAGES` 10 ta statusning hammasini aynan bir
    # marta qamraganini tekshiradi.
    assert sorted(seen) == sorted(everything)
    assert len(seen) == len(set(seen))


@pytest.mark.smoke
def test_export_combines_queue_with_panel_filters(client, admin_headers, db_session):
    """Navbat + filtr paneli birga: frontend ikkisini bitta URL'ga
    qo'shib yuboradi (`filterParams()`), eksport ham ikkisini ham
    hisobga olishi kerak."""
    tickets = _seed_one_per_queue(client, admin_headers, db_session)

    only_one = _export_tickets(client, admin_headers, unassigned=True, q=tickets["stuck_ai"])
    assert only_one == [tickets["stuck_ai"]]

    # Bir-birini kesib tashlaydigan juftlik — bo'sh fayl (sarlavha bilan).
    assert _export_tickets(client, admin_headers, unassigned=True, mine=True) == []
